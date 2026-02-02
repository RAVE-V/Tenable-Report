"""Import server-to-application mappings from Excel"""

import pandas as pd
from pathlib import Path
from typing import Dict, List
import click

from src.database.models import Server, Application, ServerApplicationMap, ConfidenceLevel
from src.database.session import get_db_session


def safe_echo(msg, err=False):
    """Echo message with fallback for Windows encoding issues"""
    emoji_map = {
        "✓": "[OK]", "✗": "[ERR]", "⚠️": "[WARN]", "⚠": "[WARN]",
        "📋": "[EXPORT]", "📂": "[FILE]", "📊": "[STATS]", "📝": "[NOTE]",
        "✨": "[*]", "💾": "[SAVE]", "⏱️": "[TIME]", "ℹ️": "[INFO]",
        "↻": "[UPD]", "🔍": "[DRY]"
    }
    try:
        click.echo(msg, err=err)
    except UnicodeEncodeError:
        safe_msg = msg
        for emoji, replacement in emoji_map.items():
            safe_msg = safe_msg.replace(emoji, replacement)
        click.echo(safe_msg, err=err)


class MappingImporter:
    """Import server-application mappings from Excel files"""
    
    REQUIRED_COLUMNS = ['server_name', 'application_name']
    OPTIONAL_COLUMNS = ['confidence', 'source', 'updated_by']
    
    def __init__(self):
        pass
    
    def validate_excel(self, df: pd.DataFrame) -> List[str]:
        """Validate Excel structure and return list of errors"""
        errors = []
        
        # Check required columns exist
        missing_cols = set(self.REQUIRED_COLUMNS) - set(df.columns)
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
        
        # Check for empty required columns (only server_name is strictly required to have a value)
        if not errors:
            # We strictly require server_name, but application_name can be empty (we'll just skip those rows)
            strict_cols = ['server_name']
            for col in strict_cols:
                if col in df.columns and df[col].isna().any():
                    empty_rows = df[df[col].isna()].index.tolist()
                    if len(empty_rows) > 10:
                        errors.append(f"Column '{col}' has empty values in rows: {', '.join(map(str, empty_rows[:10]))} and {len(empty_rows) - 10} more")
                    else:
                        errors.append(f"Column '{col}' has empty values in rows: {empty_rows}")
        
        return errors
    
    def import_from_excel(self, excel_path: Path, dry_run: bool = False) -> Dict:
        """
        Import mappings from Excel file
        
        Args:
            excel_path: Path to Excel file
            dry_run: If True, validate but don't save to database
            
        Returns:
            Dict with import statistics
        """
        # Read Excel
        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {e}")
        
        # Validate
        errors = self.validate_excel(df)
        if errors:
            raise ValueError("Excel validation failed:\n" + "\n".join(f"  - {err}" for err in errors))
            
        # Filter out rows with empty application_name
        if 'application_name' in df.columns:
            # Normalize to string and strip
            df['application_name'] = df['application_name'].fillna('').astype(str).str.strip()
            # Count total before
            total_rows = len(df)
            # Filter
            df = df[df['application_name'] != '']
            skipped_rows = total_rows - len(df)
            
            if skipped_rows > 0:
                safe_echo(f"[INFO] Skipping {skipped_rows} rows with no application name (unmapped)")
                
        if len(df) == 0:
            safe_echo("[WARN] No valid mappings found to import after filtering.")
            return {
                'total_rows': 0,
                'servers_created': 0,
                'servers_found': 0,
                'apps_created': 0,
                'apps_found': 0,
                'mappings_created': 0,
                'mappings_updated': 0,
                'errors': []
            }
        
        stats = {
            'total_rows': len(df),
            'servers_created': 0,
            'servers_found': 0,
            'apps_created': 0,
            'apps_found': 0,
            'mappings_created': 0,
            'mappings_updated': 0,
            'errors': []
        }
        
        if dry_run:
            safe_echo("[DRY] DRY RUN - No changes will be saved to database")
        
        with get_db_session() as session:
            for idx, row in df.iterrows():
                try:
                    server_name = str(row['server_name']).strip()
                    app_name = str(row['application_name']).strip()
                    
                    # Get or create Server (Case-insensitive lookup)
                    server = session.query(Server).filter(
                        (Server.hostname.ilike(server_name)) | (Server.asset_uuid == server_name)
                    ).first()
                    
                    if not server:
                        # Create new server
                        server = Server(hostname=server_name)
                        session.add(server)
                        session.flush()  # Get server_id
                        stats['servers_created'] += 1
                        safe_echo(f"  [OK] Created server: {server_name}")
                    else:
                        stats['servers_found'] += 1
                    
                    # Get or create Application (Case-insensitive lookup)
                    app = session.query(Application).filter(Application.app_name.ilike(app_name)).first()
                    
                    if not app:
                        # Create new application
                        app = Application(app_name=app_name)
                        session.add(app)
                        session.flush()  # Get app_id
                        stats['apps_created'] += 1
                        safe_echo(f"  [OK] Created application: {app_name}")
                    else:
                        stats['apps_found'] += 1
                    
                    # Check if mapping exists
                    mapping = session.query(ServerApplicationMap).filter_by(
                        server_id=server.server_id,
                        app_id=app.app_id
                    ).first()
                    
                    # Parse optional fields
                    confidence = ConfidenceLevel.MANUAL  # Default for manual imports
                    if 'confidence' in row and pd.notna(row['confidence']):
                        conf_str = str(row['confidence']).upper()
                        if conf_str in ['HIGH', 'MEDIUM', 'LOW', 'AUTO', 'MANUAL']:
                            confidence = ConfidenceLevel[conf_str]
                    
                    source = 'excel_import'
                    if 'source' in row and pd.notna(row['source']):
                        source = str(row['source']).strip()
                    
                    updated_by = None
                    if 'updated_by' in row and pd.notna(row['updated_by']):
                        updated_by = str(row['updated_by']).strip()
                    
                    if mapping:
                        # Update existing mapping
                        mapping.confidence = confidence
                        mapping.source = source
                        mapping.updated_by = updated_by
                        stats['mappings_updated'] += 1
                        safe_echo(f"  [UPD] Updated mapping: {server_name} -> {app_name}")
                    else:
                        # Create new mapping
                        mapping = ServerApplicationMap(
                            server_id=server.server_id,
                            app_id=app.app_id,
                            confidence=confidence,
                            source=source,
                            updated_by=updated_by
                        )
                        session.add(mapping)
                        stats['mappings_created'] += 1
                        safe_echo(f"  [OK] Created mapping: {server_name} -> {app_name}")
                
                except Exception as e:
                    error_msg = f"Row {idx + 2}: {str(e)}"
                    stats['errors'].append(error_msg)
                    safe_echo(f"  [ERR] Error on row {idx + 2}: {e}", err=True)
            
            if not dry_run:
                session.commit()
                safe_echo("\n[OK] Changes committed to database")
            else:
                session.rollback()  # Explicitly rollback for dry run to override context manager commit
                safe_echo("\n[DRY] DRY RUN - No changes saved")
        
        return stats
    
    def export_template(self, output_path: Path, include_existing: bool = True, servers_only: bool = False):
        """Export an Excel template with actual servers and existing mappings
        
        Args:
            output_path: Path for output Excel file
            include_existing: If True, include existing mappings from database
            servers_only: If True, only include assets with device_type='server'
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Get existing servers and mappings from database
        servers_data = []
        
        with get_db_session() as session:
            # Get all servers
            servers = session.query(Server).all()
            
            for server in servers:
                # Check for existing mappings
                mappings = session.query(ServerApplicationMap).filter_by(
                    server_id=server.server_id
                ).all()
                
                if mappings and include_existing:
                    # Server has existing mappings
                    for mapping in mappings:
                        app = session.query(Application).filter_by(
                            app_id=mapping.app_id
                        ).first()
                        servers_data.append({
                            'server_name': server.hostname,
                            'application_name': app.app_name if app else '',
                            'confidence': mapping.confidence.name if mapping.confidence else 'MANUAL',
                            'source': mapping.source or 'database',
                            'updated_by': mapping.updated_by or ''
                        })
                else:
                    # Server has no mapping - add with empty application
                    servers_data.append({
                        'server_name': server.hostname,
                        'application_name': '',
                        'confidence': 'MANUAL',
                        'source': '',
                        'updated_by': ''
                    })
            
            # ---------------------------------------------------------
            # NEW LOGIC: Also fetch assets from Vulnerabilities table
            # ---------------------------------------------------------
            from src.database.models import Vulnerability
            
            # Get distinct assets from Vulnerability table
            vuln_query = session.query(
                Vulnerability.asset_uuid, 
                Vulnerability.hostname,
                Vulnerability.device_type
            ).distinct()
            
            # Filter by device_type if servers_only is True
            if servers_only:
                vuln_query = vuln_query.filter(Vulnerability.device_type == 'server')
                safe_echo("  Filter: device_type = server")
            
            vuln_assets = vuln_query.all()
            
            # Create set of existing server hostnames/uuids for fast lookup
            existing_hostnames = {s['server_name'] for s in servers_data if s['server_name']}
            
            new_assets_count = 0
            for v_asset_uuid, v_hostname, v_device_type in vuln_assets:
                # Determine primary identifier (hostname preferred)
                identifier = v_hostname if v_hostname else v_asset_uuid
                
                if identifier and identifier not in existing_hostnames:
                    # Found a new asset not in Server inventory!
                    servers_data.append({
                        'server_name': identifier,
                        'application_name': '', # Empty for user to fill
                        'confidence': 'MANUAL',
                        'source': 'Tenable Scan (Unmapped)',
                        'updated_by': ''
                    })
                    existing_hostnames.add(identifier)
                    new_assets_count += 1
            
            if new_assets_count > 0:
                safe_echo(f"  [OK] Added {new_assets_count} unmapped assets from recent scan")

        
        # If no servers in DB, create example template
        if not servers_data:
            servers_data = [
                {
                    'server_name': 'web-prod-01.company.com',
                    'application_name': 'Payment Gateway',
                    'confidence': 'MANUAL',
                    'source': 'IT Team',
                    'updated_by': 'john.doe'
                },
                {
                    'server_name': 'db-prod-02.company.com',
                    'application_name': '',
                    'confidence': 'MANUAL',
                    'source': '',
                    'updated_by': ''
                }
            ]
            safe_echo("[INFO] No servers in database. Creating example template.")
        
        # Create DataFrame
        df = pd.DataFrame(servers_data)
        
        # Sort: unmapped servers first (easier to fill), then by server name
        df['has_mapping'] = df['application_name'].apply(lambda x: 0 if x else 1)
        df = df.sort_values(['has_mapping', 'server_name'], ascending=[False, True])
        df = df.drop('has_mapping', axis=1)
        
        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Server-App Mappings')
            
            # Access the worksheet to add formatting
            ws = writer.sheets['Server-App Mappings']
            
            # Style header row
            header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Highlight rows without application mapping (to be filled)
            highlight_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                # Column B is application_name
                if not row[1].value:  # No application name
                    for cell in row:
                        cell.fill = highlight_fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Count stats
        total = len(df)
        mapped = len(df[df['application_name'] != ''])
        unmapped = total - mapped
        
        safe_echo(f"[OK] Template exported to: {output_path}")
        safe_echo(f"   Total servers: {total}")
        safe_echo(f"   Already mapped: {mapped} (existing mappings preserved)")
        safe_echo(f"   Needs mapping: {unmapped} (highlighted in yellow)")
        safe_echo(f"\n[NOTE] Fill in the 'application_name' column for highlighted rows, then import:")
        safe_echo(f"   python -m src.cli import-mappings {output_path}")

    def export_apps_template(self, output_path: Path):
        """Export an Excel template for managing applications catalog
        
        Args:
            output_path: Path for output Excel file
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        apps_data = []
        
        with get_db_session() as session:
            # Get all applications from database
            apps = session.query(Application).all()
            
            for app in apps:
                apps_data.append({
                    'application_name': app.app_name,
                    'app_type': app.app_type or '',
                    'description': app.description or '',
                    'system_owner': app.system_owner or '',
                    'owner_team': app.owner_team or ''
                })
            
            # Also get unique app names from mappings that might not be in apps table
            from src.database.models import ServerApplicationMap
            mapped_app_ids = session.query(ServerApplicationMap.app_id).distinct().all()
            existing_app_names = {a['application_name'] for a in apps_data}
            
            for (app_id,) in mapped_app_ids:
                app = session.query(Application).filter_by(app_id=app_id).first()
                if app and app.app_name not in existing_app_names:
                    apps_data.append({
                        'application_name': app.app_name,
                        'app_type': app.app_type or '',
                        'description': app.description or '',
                        'system_owner': app.system_owner or '',
                        'owner_team': app.owner_team or ''
                    })
        
        # If no apps in DB, create example template
        if not apps_data:
            apps_data = [
                {
                    'application_name': 'Example Application',
                    'app_type': 'Web Application',
                    'description': 'Customer-facing web portal',
                    'system_owner': 'john.doe@company.com',
                    'owner_team': 'Platform Team'
                },
                {
                    'application_name': 'Payment Gateway',
                    'app_type': 'Backend Service',
                    'description': '',
                    'system_owner': '',
                    'owner_team': ''
                }
            ]
            safe_echo("[INFO] No applications in database. Creating example template.")
        
        # Create DataFrame
        df = pd.DataFrame(apps_data)
        
        # Sort by app name
        df = df.sort_values('application_name')
        
        # Write to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Applications')
            
            # Access the worksheet to add formatting
            ws = writer.sheets['Applications']
            
            # Style header row
            header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Highlight rows without owner_team (to be filled)
            highlight_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                # Column E is owner_team (index 4)
                if not row[4].value:  # No owner_team
                    for cell in row:
                        cell.fill = highlight_fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Count stats
        total = len(df)
        with_team = len(df[df['owner_team'] != ''])
        needs_team = total - with_team
        
        safe_echo(f"[OK] Applications template exported to: {output_path}")
        safe_echo(f"   Total applications: {total}")
        safe_echo(f"   With owner_team: {with_team}")
        safe_echo(f"   Needs owner_team: {needs_team} (highlighted in yellow)")
        safe_echo(f"\n[NOTE] Fill in 'system_owner' and 'owner_team' columns, then import:")
        safe_echo(f"   python -m src.cli import-mappings {output_path} --type apps")

    def import_apps_from_excel(self, excel_path: Path, dry_run: bool = False) -> Dict:
        """
        Import application metadata from Excel file
        
        Args:
            excel_path: Path to Excel file
            dry_run: If True, validate but don't save to database
            
        Returns:
            Dict with import statistics
        """
        # Read Excel
        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {e}")
        
        # Validate required column
        if 'application_name' not in df.columns:
            raise ValueError("Excel must contain 'application_name' column")
        
        # Filter out rows with empty application_name
        df['application_name'] = df['application_name'].fillna('').astype(str).str.strip()
        df = df[df['application_name'] != '']
        
        if len(df) == 0:
            safe_echo("[WARN] No valid applications found to import.")
            return {
                'total_rows': 0,
                'apps_created': 0,
                'apps_updated': 0,
                'errors': []
            }
        
        stats = {
            'total_rows': len(df),
            'apps_created': 0,
            'apps_updated': 0,
            'errors': []
        }
        
        if dry_run:
            safe_echo("[DRY] DRY RUN - No changes will be saved to database")
        
        with get_db_session() as session:
            for idx, row in df.iterrows():
                try:
                    app_name = str(row['application_name']).strip()
                    
                    # Get optional fields
                    app_type = str(row.get('app_type', '')).strip() if pd.notna(row.get('app_type')) else None
                    description = str(row.get('description', '')).strip() if pd.notna(row.get('description')) else None
                    system_owner = str(row.get('system_owner', '')).strip() if pd.notna(row.get('system_owner')) else None
                    owner_team = str(row.get('owner_team', '')).strip() if pd.notna(row.get('owner_team')) else None
                    
                    # Find existing application (case-insensitive)
                    app = session.query(Application).filter(Application.app_name.ilike(app_name)).first()
                    
                    if app:
                        # Update existing application
                        if app_type:
                            app.app_type = app_type
                        if description:
                            app.description = description
                        if system_owner:
                            app.system_owner = system_owner
                        if owner_team:
                            app.owner_team = owner_team
                        stats['apps_updated'] += 1
                        safe_echo(f"  [UPD] Updated: {app_name}")
                    else:
                        # Create new application
                        app = Application(
                            app_name=app_name,
                            app_type=app_type,
                            description=description,
                            system_owner=system_owner,
                            owner_team=owner_team
                        )
                        session.add(app)
                        stats['apps_created'] += 1
                        safe_echo(f"  [OK] Created: {app_name}")
                
                except Exception as e:
                    error_msg = f"Row {idx + 2}: {str(e)}"
                    stats['errors'].append(error_msg)
                    safe_echo(f"  [ERR] Error on row {idx + 2}: {e}", err=True)
            
            if not dry_run:
                session.commit()
                safe_echo("\n[OK] Changes committed to database")
            else:
                session.rollback()
                safe_echo("\n[DRY] DRY RUN - No changes saved")
        
        return stats

