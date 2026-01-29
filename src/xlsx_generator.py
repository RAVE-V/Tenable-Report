"""XLSX report generator"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class XLSXReportGenerator:
    """Generate XLSX reports from vulnerability data"""
    
    SEVERITY_COLORS = {
        "Critical": "FFE74C3C",  # Red
        "High": "FFE67E22",      # Orange
        "Medium": "FFF39C12",    # Yellow
        "Low": "FF3498DB",       # Blue
        "Info": "FF95A5A6"       # Gray
    }
    
    def generate(self, vulnerabilities: List[Dict], output_path: Path, metadata: Dict = None):
        """
        Generate XLSX report
        
        Args:
            vulnerabilities: List of normalized vulnerability dictionaries
            output_path: Path to save XLSX file
            metadata: Optional metadata (filters, timestamp, etc.)
        """
        logger.info(f"Generating XLSX report with {len(vulnerabilities)} vulnerabilities")
        
        # Convert to DataFrame
        df = pd.DataFrame(vulnerabilities)
        
        # Select and order columns (matching reference implementation)
        columns = [
            "asset_uuid",
            "hostname",
            "ipv4",
            "operating_system",
            "plugin_id",
            "plugin_name",
            "description",
            "cve",
            "solution",
            "synopsis",
            "see_also",
            "exploit_available",
            "has_patch",
            "severity",
            "state",
            "first_found",
            "last_found"
        ]
        
        # Ensure all columns exist
        for col in columns:
            if col not in df.columns:
                df[col] = None
        
        df = df[columns]
        
        # Format CVE lists as strings
        if "cve" in df.columns:
            df["cve"] = df["cve"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
        
        # Format see_also lists as strings
        if "see_also" in df.columns:
            df["see_also"] = df["see_also"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
        
        # Create Excel file with formatting
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Vulnerabilities", index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets["Vulnerabilities"]
            
            # Format header row
            header_fill = PatternFill(start_color="FF34495E", end_color="FF34495E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFFFF")
            
            for col_num in range(1, len(columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Format severity cells with color coding
            if "severity" in df.columns:
                severity_col = columns.index("severity") + 1
                for row_num in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_num, column=severity_col)
                    severity = cell.value
                    if severity in self.SEVERITY_COLORS:
                        cell.fill = PatternFill(
                            start_color=self.SEVERITY_COLORS[severity],
                            end_color=self.SEVERITY_COLORS[severity],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFFFF")
            
            # Auto-adjust column widths
            for col_num, column in enumerate(columns, start=1):
                col_letter = get_column_letter(col_num)
                
                # Calculate max width based on column name and content
                max_width = len(str(column)) + 2
                for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=col_num, max_col=col_num):
                    cell_value = str(row[0].value) if row[0].value else ""
                    max_width = max(max_width, min(len(cell_value), 50))  # Cap at 50
                
                worksheet.column_dimensions[col_letter].width = max_width
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
            
            # Add metadata sheet if provided
            if metadata:
                self._add_metadata_sheet(workbook, metadata)
        
        logger.info(f"XLSX report saved to {output_path}")
    
    def _add_metadata_sheet(self, workbook: Workbook, metadata: Dict):
        """Add metadata sheet to workbook"""
        ws = workbook.create_sheet("Report Metadata")
        
        # Add metadata rows
        ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["Filters", str(metadata.get("filters", {}))])
        ws.append(["Total Vulnerabilities", metadata.get("total_vulns", 0)])
        ws.append(["Total Assets", metadata.get("total_assets", 0)])
        ws.append(["Runtime (seconds)", metadata.get("runtime_seconds", 0)])
        
        # Format
        for row in ws.iter_rows(min_row=1, max_row=5):
            row[0].font = Font(bold=True)
            row[0].fill = PatternFill(start_color="FFECF0F1", end_color="FFECF0F1", fill_type="solid")
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60
    
    def generate_server_report(self, output_path: Path, servers: List[tuple], stats: Dict, metadata: Dict = None):
        """
        Generate server-focused XLSX report
        
        Args:
            output_path: Path to save XLSX file
            servers: List of (hostname, server_data) tuples sorted by severity
            stats: Overall statistics dictionary
            metadata: Optional metadata
        """
        logger.info(f"Generating server-focused XLSX report with {len(servers)} servers")
        
        # Create workbook
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Sheet 1: Server Summary
            summary_data = []
            for hostname, server_data in servers:
                summary_data.append({
                    "Hostname": hostname,
                    "IP Address": server_data.get("ipv4", "N/A"),
                    "Operating System": server_data.get("os", "N/A"),
                    "Total Vulnerabilities": server_data["total_vulns"],
                    "Critical": server_data["severity_counts"]["critical"],
                    "High": server_data["severity_counts"]["high"],
                    "Medium": server_data["severity_counts"]["medium"],
                    "Low": server_data["severity_counts"]["low"],
                    "Quick Wins": server_data["quick_wins"]
                })
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Server Summary", index=False)
            
            # Format Server Summary sheet
            ws_summary = writer.sheets["Server Summary"]
            self._format_server_summary_sheet(ws_summary, len(summary_data))
            
            # Sheet 2: All Vulnerabilities (flat list)
            all_vulns = []
            for hostname, server_data in servers:
                for vuln in server_data["vulnerabilities"]:
                    all_vulns.append(vuln)
            
            if all_vulns:
                vulns_df = pd.DataFrame(all_vulns)
                
                # Select columns
                columns = [
                    "hostname", "ipv4", "operating_system",
                    "plugin_id", "plugin_name", "severity", "state",
                    "cve", "solution", "has_patch", "exploit_available",
                    "first_found", "last_found"
                ]
                
                # Ensure all columns exist
                for col in columns:
                    if col not in vulns_df.columns:
                        vulns_df[col] = None
                
                vulns_df = vulns_df[columns]
                
                # Format lists as strings
                if "cve" in vulns_df.columns:
                    vulns_df["cve"] = vulns_df["cve"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
                
                vulns_df.to_excel(writer, sheet_name="All Vulnerabilities", index=False)
                
                # Format All Vulnerabilities sheet
                ws_vulns = writer.sheets["All Vulnerabilities"]
                self._format_vulnerabilities_sheet(ws_vulns, vulns_df, columns)
            
            # Sheet 3: Statistics
            stats_data = [
                ["Total Servers", stats["total_servers"]],
                ["Total Vulnerabilities", stats["total_vulns"]],
                ["Total Quick Wins", stats["total_quick_wins"]],
                ["", ""],
                ["Severity Breakdown", ""],
                ["Critical", stats["severity_totals"]["critical"]],
                ["High", stats["severity_totals"]["high"]],
                ["Medium", stats["severity_totals"]["medium"]],
                ["Low", stats["severity_totals"]["low"]]
            ]
            
            stats_df = pd.DataFrame(stats_data, columns=["Metric", "Value"])
            stats_df.to_excel(writer, sheet_name="Statistics", index=False)
            
            # Format Statistics sheet
            ws_stats = writer.sheets["Statistics"]
            self._format_stats_sheet(ws_stats)
            
            # Add metadata sheet if provided
            if metadata:
                self._add_metadata_sheet(writer.book, metadata)
        
        logger.info(f"Server report saved to {output_path}")
    
    def _format_server_summary_sheet(self, ws, row_count):
        """Format the Server Summary sheet"""
        # Header formatting
        header_fill = PatternFill(start_color="FF34495E", end_color="FF34495E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFFFF")
        
        for col in range(1, 10):  # 9 columns
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color code severity columns
        for row in range(2, row_count + 2):
            # Critical
            cell = ws.cell(row=row, column=5)
            if cell.value and cell.value > 0:
                cell.fill = PatternFill(start_color=self.SEVERITY_COLORS["Critical"], end_color=self.SEVERITY_COLORS["Critical"], fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFFFF")
            
            # High
            cell = ws.cell(row=row, column=6)
            if cell.value and cell.value > 0:
                cell.fill = PatternFill(start_color=self.SEVERITY_COLORS["High"], end_color=self.SEVERITY_COLORS["High"], fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFFFF")
            
            # Medium
            cell = ws.cell(row=row, column=7)
            if cell.value and cell.value > 0:
                cell.fill = PatternFill(start_color=self.SEVERITY_COLORS["Medium"], end_color=self.SEVERITY_COLORS["Medium"], fill_type="solid")
                cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 35  # Hostname
        ws.column_dimensions["B"].width = 18  # IP
        ws.column_dimensions["C"].width = 30  # OS
        ws.column_dimensions["D"].width = 12  # Total
        ws.column_dimensions["E"].width = 10  # Critical
        ws.column_dimensions["F"].width = 10  # High
        ws.column_dimensions["G"].width = 10  # Medium
        ws.column_dimensions["H"].width = 10  # Low
        ws.column_dimensions["I"].width = 12  # Quick Wins
        
        # Freeze header
        ws.freeze_panes = "A2"
    
    def _format_vulnerabilities_sheet(self, ws, df, columns):
        """Format the All Vulnerabilities sheet"""
        # Header formatting
        header_fill = PatternFill(start_color="FF34495E", end_color="FF34495E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFFFF")
        
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color code severity column
        if "severity" in columns:
            severity_col = columns.index("severity") + 1
            for row_num in range(2, len(df) + 2):
                cell = ws.cell(row=row_num, column=severity_col)
                severity = cell.value
                if severity in self.SEVERITY_COLORS:
                    cell.fill = PatternFill(
                        start_color=self.SEVERITY_COLORS[severity],
                        end_color=self.SEVERITY_COLORS[severity],
                        fill_type="solid"
                    )
                    cell.font = Font(bold=True, color="FFFFFFFF")
        
        # Auto-adjust column widths
        for col_num, column in enumerate(columns, start=1):
            col_letter = get_column_letter(col_num)
            max_width = len(str(column)) + 2
            for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=col_num, max_col=col_num):
                cell_value = str(row[0].value) if row[0].value else ""
                max_width = max(max_width, min(len(cell_value), 50))
            ws.column_dimensions[col_letter].width = max_width
        
        # Freeze header
        ws.freeze_panes = "A2"
    
    def _format_stats_sheet(self, ws):
        """Format the Statistics sheet"""
        # Bold the metric column
        for row in range(1, 11):
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True)
            
            # Highlight section headers
            if row in [1, 5]:
                cell.fill = PatternFill(start_color="FFECF0F1", end_color="FFECF0F1", fill_type="solid")
        
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

